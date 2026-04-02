"""
Módulo para gestionar registros de ingreso de peatones.
Incluye creación de carpetas, generación de tickets y almacenamiento de imágenes en Excel.
"""

import base64
import openpyxl
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


class RegistroPeatonal:
    """
    Gestiona el registro de ingreso de peatones.
    Organiza carpetas por fecha (YEAR/MM_MonthName/DD/TICKET_#####)
    y mantiene un Excel maestro con información de registros.
    """
    
    MESES = {
        1: "01_Enero",
        2: "02_Febrero",
        3: "03_Marzo",
        4: "04_Abril",
        5: "05_Mayo",
        6: "06_Junio",
        7: "07_Julio",
        8: "08_Agosto",
        9: "09_Septiembre",
        10: "10_Octubre",
        11: "11_Noviembre",
        12: "12_Diciembre"
    }
    
    def __init__(self, ruta_base: str):
        """
        Inicializa el gestor de registros de peatones.
        
        Args:
            ruta_base: Ruta base donde se crearán los registros.
                      Ej: C:\\Users\\LENOVO\\Documents\\Base de datos\\DMT_Gestion_Ingreso\\Ingreso Peatonal
        """
        self.ruta_base = Path(ruta_base)
        self.ruta_base.mkdir(parents=True, exist_ok=True)
        
        self.año_actual = datetime.now().year
        self.ruta_excel = self.ruta_base / "registro_historico_peatones.xlsx"
        
        # Crear estructura base y Excel
        self._crear_estructura_base()
        self._inicializar_excel()
    
    def _crear_estructura_base(self):
        """Crea la carpeta del año si no existe"""
        ruta_año = self.ruta_base / str(self.año_actual)
        ruta_año.mkdir(parents=True, exist_ok=True)
    
    def _inicializar_excel(self):
        """
        Inicializa el archivo Excel si no existe.
        Si ya existe, verifica que tenga la estructura correcta.
        """
        if self.ruta_excel.exists():
            # Verificar que el Excel tiene la hoja correcta
            try:
                wb = openpyxl.load_workbook(self.ruta_excel)
                if "Registros" not in wb.sheetnames:
                    wb.close()
                    self._crear_excel_nuevo()
                else:
                    wb.close()
            except Exception as e:
                print(f"Error verificando Excel: {e}")
                self._crear_excel_nuevo()
        else:
            self._crear_excel_nuevo()
    
    def _crear_excel_nuevo(self):
        """Crea un nuevo archivo Excel con la estructura de registros de peatones"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros"
        
        # Encabezados: TICKET, PERSONA, CÉDULA, DEPARTAMENTO, INGRESO, SALIDA/ESTADO, FECHA_REGISTRO
        encabezados = [
            "TICKET",
            "PERSONA",
            "CÉDULA",
            "DEPARTAMENTO",
            "INGRESO",
            "SALIDA/ESTADO",
            "FECHA_REGISTRO"
        ]
        
        # Insertar encabezados con estilos
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=encabezado)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15  # TICKET
        ws.column_dimensions['B'].width = 30  # PERSONA
        ws.column_dimensions['C'].width = 15  # CÉDULA
        ws.column_dimensions['D'].width = 20  # DEPARTAMENTO
        ws.column_dimensions['E'].width = 12  # INGRESO
        ws.column_dimensions['F'].width = 12  # SALIDA/ESTADO
        ws.column_dimensions['G'].width = 20  # FECHA_REGISTRO
        
        wb.save(self.ruta_excel)
        wb.close()
    
    def _obtener_siguiente_ticket(self) -> int:
        """Obtiene el siguiente número de ticket"""
        wb = openpyxl.load_workbook(self.ruta_excel)
        ws = wb["Registros"]
        
        # Contar filas con datos (excluyendo encabezado)
        filas_con_datos = ws.max_row - 1  # Restar 1 por el encabezado
        siguiente_numero = filas_con_datos + 1
        
        wb.close()
        return siguiente_numero
    
    def _crear_carpeta_ticket(self, numero_ticket: int) -> Path:
        """
        Crea la estructura de carpetas para un ticket
        
        Args:
            numero_ticket: Número del ticket
        
        Returns:
            Path: Ruta de la carpeta del ticket
        """
        ahora = datetime.now()
        mes_nombre = self.MESES[ahora.month]
        día = f"{ahora.day:02d}"
        
        # Construir ruta
        ruta_ticket = (
            self.ruta_base / 
            str(self.año_actual) / 
            mes_nombre / 
            día / 
            f"TICKET_{numero_ticket:06d}"
        )
        
        ruta_ticket.mkdir(parents=True, exist_ok=True)
        
        return ruta_ticket
    
    def _guardar_imagen(self, ruta_ticket: Path, imagen_fuente, nombre_archivo: str) -> str:
        """
        Guarda una imagen desde base64 o bytes
        
        Args:
            ruta_ticket: Ruta de la carpeta del ticket
            imagen_fuente: Imagen codificada en base64 (str) o bytes (bytes)
            nombre_archivo: Nombre del archivo (ej: cedula.jpg)
        
        Returns:
            str: Ruta del archivo guardado
        """
        try:
            # Decodificar si es base64, de lo contrario asumir que es bytes
            if isinstance(imagen_fuente, str):
                imagen_bytes = base64.b64decode(imagen_fuente)
            else:
                imagen_bytes = imagen_fuente
            
            # Construir ruta de archivo
            ruta_archivo = ruta_ticket / nombre_archivo
            
            # Guardar archivo
            with open(ruta_archivo, 'wb') as f:
                f.write(imagen_bytes)
            
            return str(ruta_archivo)
        
        except Exception as e:
            print(f"Error guardando imagen {nombre_archivo}: {e}")
            return None
    
    def guardar_registro(self, datos: Dict) -> Dict:
        """
        Guarda un registro completo de ingreso de peatón
        
        Args:
            datos: Diccionario con:
                {
                    'persona': str,  # Nombres y apellidos
                    'cedula': str,
                    'departamento': str,
                    'imagen_cedula_base64': str o bytes (opcional),
                    'imagen_usuario_base64': str o bytes (opcional),
                    'hora_ingreso': str (opcional, formato HH:MM:SS)
                }
        
        Returns:
            dict con resultado y detalles del registro
        """
        try:
            # Obtener siguiente número de ticket
            numero_ticket = self._obtener_siguiente_ticket()
            
            # Crear carpeta del ticket
            ruta_ticket = self._crear_carpeta_ticket(numero_ticket)
            
            # Guardar imágenes
            rutas_imagenes = {}
            if datos.get('imagen_cedula_base64'):
                rutas_imagenes['cedula'] = self._guardar_imagen(
                    ruta_ticket, 
                    datos['imagen_cedula_base64'], 
                    'cedula.jpg'
                )
            
            if datos.get('imagen_usuario_base64'):
                rutas_imagenes['usuario'] = self._guardar_imagen(
                    ruta_ticket,
                    datos['imagen_usuario_base64'],
                    'usuario.jpg'
                )
            
            # Agregar datos al Excel
            self._agregar_fila_excel(
                numero_ticket,
                datos['persona'],
                datos['cedula'],
                datos['departamento'],
                datos.get('hora_ingreso', datetime.now().strftime("%H:%M:%S")),
                ''  # hora_salida vacía al inicio
            )
            
            return {
                'success': True,
                'numero_ticket': numero_ticket,
                'ruta_ticket': str(ruta_ticket),
                'imagenes': rutas_imagenes,
                'mensaje': f'Registro TICKET_{numero_ticket:06d} guardado exitosamente'
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'numero_ticket': None
            }
    
    def _agregar_fila_excel(self, numero_ticket: int, persona: str,
                           cedula: str, departamento: str, hora_ingreso: str, 
                           hora_salida: str):
        """Agrega una fila al Excel"""
        wb = openpyxl.load_workbook(self.ruta_excel)
        ws = wb["Registros"]
        
        # Obtener siguiente fila
        siguiente_fila = ws.max_row + 1
        
        # Datos a insertar: TICKET, PERSONA, CÉDULA, DEPARTAMENTO, INGRESO, SALIDA/ESTADO, FECHA_REGISTRO
        datos = [
            f"TICKET_{numero_ticket:06d}",
            persona,
            cedula,
            departamento,
            hora_ingreso,
            hora_salida,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        # Insertar datos
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=siguiente_fila, column=col, value=valor)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        wb.save(self.ruta_excel)
        wb.close()
    
    def actualizar_hora_salida(self, numero_ticket: int, hora_salida: str) -> bool:
        """
        Actualiza la hora de salida de un ticket
        
        Args:
            numero_ticket: Número del ticket
            hora_salida: Hora de salida (formato HH:MM:SS)
        
        Returns:
            bool: True si se actualizó, False si no encontró
        """
        try:
            wb = openpyxl.load_workbook(self.ruta_excel)
            ws = wb["Registros"]
            
            # Buscar el ticket
            for fila in range(2, ws.max_row + 1):
                numero_ticket_celda = ws.cell(row=fila, column=1).value
                if numero_ticket_celda == f"TICKET_{numero_ticket:06d}":
                    # Actualizar columna de SALIDA/ESTADO (columna 6)
                    ws.cell(row=fila, column=6, value=hora_salida)
                    wb.save(self.ruta_excel)
                    wb.close()
                    return True
            
            wb.close()
            return False
        
        except Exception as e:
            print(f"Error actualizando hora de salida: {e}")
            return False
    
    def actualizar_hora_salida_por_ticket(self, ticket_ref: str, hora_salida: Optional[str] = None) -> Dict:
        """
        Actualiza la hora de salida por número o código de ticket.
        
        Args:
            ticket_ref: Número (ej: 1) o código (ej: TICKET_000001)
            hora_salida: Hora de salida HH:MM:SS. Si no se envía, usa la hora actual.
        
        Returns:
            dict con resultado de la actualización.
        """
        try:
            codigo_ticket = self._normalizar_codigo_ticket(ticket_ref)
            if not codigo_ticket:
                return {
                    'success': False,
                    'error': 'Ticket inválido'
                }
            
            hora_final = hora_salida or datetime.now().strftime("%H:%M:%S")
            
            wb = openpyxl.load_workbook(self.ruta_excel)
            ws = wb["Registros"]
            
            for fila in range(2, ws.max_row + 1):
                numero_ticket_celda = str(ws.cell(row=fila, column=1).value or "").strip().upper()
                if numero_ticket_celda == codigo_ticket:
                    ws.cell(row=fila, column=6, value=hora_final)
                    wb.save(self.ruta_excel)
                    wb.close()
                    return {
                        'success': True,
                        'ticket': codigo_ticket,
                        'hora_salida': hora_final,
                        'mensaje': f'Hora de salida actualizada para {codigo_ticket}'
                    }
            
            wb.close()
            return {
                'success': False,
                'error': f'No se encontró el ticket {codigo_ticket}'
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def _normalizar_codigo_ticket(self, ticket_ref: str) -> str:
        """Convierte referencias de ticket a formato TICKET_######."""
        if ticket_ref is None:
            return ""
        
        valor = str(ticket_ref).strip()
        if not valor:
            return ""
        
        if valor.isdigit():
            return f"TICKET_{int(valor):06d}"
        
        valor_upper = valor.upper()
        if valor_upper.startswith("TICKET_"):
            sufijo = valor_upper.split("_", 1)[1]
            if sufijo.isdigit():
                return f"TICKET_{int(sufijo):06d}"
            return valor_upper
        
        return valor_upper
    
    def obtener_todos_tickets(self) -> Dict:
        """
        Lee el Excel maestro y retorna todos los tickets registrados.
        
        Returns:
            dict con estado y lista de tickets.
        """
        try:
            wb = openpyxl.load_workbook(self.ruta_excel, data_only=True)
            ws = wb["Registros"]
            
            tickets = []
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if not fila or not fila[0]:
                    continue
                
                tickets.append({
                    'ticket': fila[0],
                    'persona': fila[1],
                    'cedula': fila[2],
                    'departamento': fila[3],
                    'ingreso': fila[4],
                    'salida_estado': fila[5],
                    'fecha_registro': fila[6]
                })
            
            wb.close()
            return {
                'success': True,
                'cantidad': len(tickets),
                'tickets': tickets
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'tickets': []
            }
    
    def obtener_fotos_ticket(self, numero_ticket: int) -> Dict:
        """
        Obtiene las fotos asociadas a un ticket.
        
        Args:
            numero_ticket: Número del ticket
        
        Returns:
            dict con rutas de imágenes
        """
        try:
            # Buscar carpeta del ticket (puede estar en cualquier mes/día)
            año_actual = datetime.now().year
            ruta_año = self.ruta_base / str(año_actual)
            
            # Buscar recursivamente
            rutas_imagenes = {}
            for ruta_mes in ruta_año.iterdir():
                if not ruta_mes.is_dir():
                    continue
                
                for ruta_día in ruta_mes.iterdir():
                    if not ruta_día.is_dir():
                        continue
                    
                    ruta_ticket_esperada = ruta_día / f"TICKET_{numero_ticket:06d}"
                    if ruta_ticket_esperada.exists():
                        # Buscar imágenes
                        for archivo in ruta_ticket_esperada.iterdir():
                            if archivo.suffix.lower() in ['.jpg', '.jpeg', '.png']:
                                rutas_imagenes[archivo.stem] = str(archivo)
                        
                        return {
                            'success': True,
                            'numero_ticket': numero_ticket,
                            'ruta_ticket': str(ruta_ticket_esperada),
                            'imagenes': rutas_imagenes
                        }
            
            return {
                'success': False,
                'error': f'Ticket TICKET_{numero_ticket:06d} no encontrado',
                'imagenes': {}
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'imagenes': {}
            }

    def _buscar_carpeta_ticket(self, codigo_ticket: str) -> Optional[Path]:
        """Busca la carpeta física asociada a un ticket."""
        if not codigo_ticket:
            return None

        rutas_raiz = [self.ruta_base / str(self.año_actual), self.ruta_base]
        for raiz in rutas_raiz:
            if not raiz.exists():
                continue

            for ruta in raiz.rglob(codigo_ticket):
                if ruta.is_dir() and ruta.name.upper() == codigo_ticket:
                    return ruta

        return None

    def obtener_fotos_por_ticket(self, ticket_ref: str) -> Dict:
        """
        Recupera imágenes asociadas a un ticket y las retorna en base64.

        Args:
            ticket_ref: Número (ej: 1) o código (ej: TICKET_000001)

        Returns:
            dict con imágenes encontradas y faltantes.
        """
        try:
            codigo_ticket = self._normalizar_codigo_ticket(ticket_ref)
            if not codigo_ticket:
                return {
                    'success': False,
                    'error': 'Ticket inválido'
                }

            carpeta_ticket = self._buscar_carpeta_ticket(codigo_ticket)
            if carpeta_ticket is None:
                return {
                    'success': False,
                    'error': f'No se encontró carpeta para el ticket {codigo_ticket}'
                }

            archivos_esperados = {
                'cedula': 'cedula.jpg',
                'usuario': 'usuario.jpg',
            }

            fotos = {}
            faltantes = []
            for clave, nombre_archivo in archivos_esperados.items():
                ruta_archivo = carpeta_ticket / nombre_archivo
                if not ruta_archivo.exists() or not ruta_archivo.is_file():
                    faltantes.append(nombre_archivo)
                    continue

                with open(ruta_archivo, 'rb') as image_file:
                    contenido = image_file.read()

                fotos[clave] = {
                    'archivo': nombre_archivo,
                    'size_bytes': len(contenido),
                    'image_base64': base64.b64encode(contenido).decode('utf-8')
                }

            return {
                'success': True,
                'ticket': codigo_ticket,
                'ruta_ticket': str(carpeta_ticket),
                'total_fotos': len(fotos),
                'faltantes': faltantes,
                'fotos': fotos,
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }


# Instancia global (se inicializará en la API)
registro_manager = None


def inicializar_registro_manager(ruta_base: str):
    """Inicializa el manager de registros peatonales con una ruta base"""
    global registro_manager
    registro_manager = RegistroPeatonal(ruta_base)
