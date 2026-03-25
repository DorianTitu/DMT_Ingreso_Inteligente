"""
Módulo de gestión de registros vehiculares
Maneja la creación de estructura de carpetas, Excel maestro y guardado de imágenes
"""

import os
import base64
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Optional


class RegistroVehicular:
    """Gestor de registros vehiculares con estructura de carpetas y Excel"""
    
    # Mapeo de meses
    MESES = {
        1: "01_Enero", 2: "02_Febrero", 3: "03_Marzo", 4: "04_Abril",
        5: "05_Mayo", 6: "06_Junio", 7: "07_Julio", 8: "08_Agosto",
        9: "09_Septiembre", 10: "10_Octubre", 11: "11_Noviembre", 12: "12_Diciembre"
    }
    
    def __init__(self, ruta_base: str):
        """
        Inicializa el gestor de registros
        
        Args:
            ruta_base: Ruta base donde guardar los registros
                      Ej: C:/Users/LENOVO/Documents/Base de datos/DMT_Gestion_Ingreso/Ingreso Vehicular
        """
        self.ruta_base = Path(ruta_base)
        self.ruta_excel = self.ruta_base / "registro_historico_vehiculos.xlsx"
        self.año_actual = datetime.now().year
        
        # Crear estructura base si no existe
        self._crear_estructura_base()
        
        # Crear o cargar Excel
        self._inicializar_excel()
    
    def _crear_estructura_base(self):
        """Crea la estructura base de carpetas"""
        self.ruta_base.mkdir(parents=True, exist_ok=True)
        (self.ruta_base / str(self.año_actual)).mkdir(exist_ok=True)
    
    def _inicializar_excel(self):
        """Crea o carga el archivo Excel maestro"""
        if not self.ruta_excel.exists():
            self._crear_excel_nuevo()
        else:
            # Verificar que tenga las columnas correctas
            wb = openpyxl.load_workbook(self.ruta_excel)
            if "Registros" not in wb.sheetnames:
                self._crear_excel_nuevo()
    
    def _crear_excel_nuevo(self):
        """Crea un Excel nuevo con estructura definida"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros"
        
        # Encabezados
        encabezados = [
            "Número de Ticket",
            "Nombres",
            "Apellidos",
            "Cédula",
            "Hora de Ingreso",
            "Hora de Salida",
            "Departamento",
            "Motivo",
            "Fecha de Registro"
        ]
        
        # Aplicar estilos a encabezados
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=encabezado)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columnas
        anchos = [18, 15, 15, 15, 16, 16, 18, 20, 16]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
        
        wb.save(self.ruta_excel)
    
    def _obtener_siguiente_ticket(self) -> int:
        """Obtiene el siguiente número de ticket"""
        wb = openpyxl.load_workbook(self.ruta_excel)
        ws = wb["Registros"]
        
        # Contar filas con datos (excluyendo encabezado)
        filas_con_datos = ws.max_row - 1  # Restar 1 por el encabezado
        siguiente_numero = filas_con_datos + 1
        
        wb.close()
        return siguiente_numero
    
    def _crear_carpeta_ticket(self, numero_ticket: int) -> Path:
        """
        Crea la estructura de carpetas para un ticket
        
        Args:
            numero_ticket: Número del ticket
        
        Returns:
            Path: Ruta de la carpeta del ticket
        """
        ahora = datetime.now()
        mes_nombre = self.MESES[ahora.month]
        día = f"{ahora.day:02d}"
        
        # Construir ruta
        ruta_ticket = (
            self.ruta_base / 
            str(self.año_actual) / 
            mes_nombre / 
            día / 
            f"TICKET_{numero_ticket:06d}"
        )
        
        ruta_ticket.mkdir(parents=True, exist_ok=True)
        
        return ruta_ticket
    
    def _guardar_imagen(self, ruta_ticket: Path, imagen_base64: str, nombre_archivo: str) -> str:
        """
        Guarda una imagen desde base64
        
        Args:
            ruta_ticket: Ruta de la carpeta del ticket
            imagen_base64: Imagen codificada en base64
            nombre_archivo: Nombre del archivo (ej: cedula.jpg)
        
        Returns:
            str: Ruta del archivo guardado
        """
        try:
            # Decodificar base64
            imagen_bytes = base64.b64decode(imagen_base64)
            
            # Construir ruta de archivo
            ruta_archivo = ruta_ticket / nombre_archivo
            
            # Guardar archivo
            with open(ruta_archivo, 'wb') as f:
                f.write(imagen_bytes)
            
            return str(ruta_archivo)
        
        except Exception as e:
            print(f"Error guardando imagen {nombre_archivo}: {e}")
            return None
    
    def guardar_registro(self, datos: Dict) -> Dict:
        """
        Guarda un registro completo de ingreso vehicular
        
        Args:
            datos: Diccionario con:
                {
                    'nombres': str,
                    'apellidos': str,
                    'cedula': str,
                    'departamento': str,
                    'motivo': str,
                    'imagen_cedula_base64': str (opcional),
                    'imagen_usuario_base64': str (opcional),
                    'imagen_placa_base64': str (opcional),
                    'hora_ingreso': str (opcional, formato HH:MM:SS)
                }
        
        Returns:
            dict con resultado y detalles del registro
        """
        try:
            # Obtener siguiente número de ticket
            numero_ticket = self._obtener_siguiente_ticket()
            
            # Crear carpeta del ticket
            ruta_ticket = self._crear_carpeta_ticket(numero_ticket)
            
            # Guardar imágenes
            rutas_imagenes = {}
            if datos.get('imagen_cedula_base64'):
                rutas_imagenes['cedula'] = self._guardar_imagen(
                    ruta_ticket, 
                    datos['imagen_cedula_base64'], 
                    'cedula.jpg'
                )
            
            if datos.get('imagen_usuario_base64'):
                rutas_imagenes['usuario'] = self._guardar_imagen(
                    ruta_ticket,
                    datos['imagen_usuario_base64'],
                    'usuario.jpg'
                )
            
            if datos.get('imagen_placa_base64'):
                rutas_imagenes['placa'] = self._guardar_imagen(
                    ruta_ticket,
                    datos['imagen_placa_base64'],
                    'placa.jpg'
                )
            
            # Agregar datos al Excel
            self._agregar_fila_excel(
                numero_ticket,
                datos['nombres'],
                datos['apellidos'],
                datos['cedula'],
                datos.get('hora_ingreso', datetime.now().strftime("%H:%M:%S")),
                datos.get('hora_salida', ''),
                datos['departamento'],
                datos['motivo']
            )
            
            return {
                'success': True,
                'numero_ticket': numero_ticket,
                'ruta_ticket': str(ruta_ticket),
                'imagenes': rutas_imagenes,
                'mensaje': f'Registro TICKET_{numero_ticket:06d} guardado exitosamente'
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'numero_ticket': None
            }
    
    def _agregar_fila_excel(self, numero_ticket: int, nombres: str, apellidos: str,
                           cedula: str, hora_ingreso: str, hora_salida: str,
                           departamento: str, motivo: str):
        """Agrega una fila al Excel"""
        wb = openpyxl.load_workbook(self.ruta_excel)
        ws = wb["Registros"]
        
        # Obtener siguiente fila
        siguiente_fila = ws.max_row + 1
        
        # Datos a insertar
        datos = [
            f"TICKET_{numero_ticket:06d}",
            nombres,
            apellidos,
            cedula,
            hora_ingreso,
            hora_salida,
            departamento,
            motivo,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        # Insertar datos
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=siguiente_fila, column=col, value=valor)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        wb.save(self.ruta_excel)
        wb.close()
    
    def actualizar_hora_salida(self, numero_ticket: int, hora_salida: str) -> bool:
        """
        Actualiza la hora de salida de un ticket
        
        Args:
            numero_ticket: Número del ticket
            hora_salida: Hora de salida (formato HH:MM:SS)
        
        Returns:
            bool: True si se actualizó, False si no encontró
        """
        try:
            wb = openpyxl.load_workbook(self.ruta_excel)
            ws = wb["Registros"]
            
            # Buscar el ticket
            for fila in range(2, ws.max_row + 1):
                numero_ticket_celda = ws.cell(row=fila, column=1).value
                if numero_ticket_celda == f"TICKET_{numero_ticket:06d}":
                    # Actualizar columna de hora de salida (columna 6)
                    ws.cell(row=fila, column=6, value=hora_salida)
                    wb.save(self.ruta_excel)
                    wb.close()
                    return True
            
            wb.close()
            return False
        
        except Exception as e:
            print(f"Error actualizando hora de salida: {e}")
            return False

    def _normalizar_codigo_ticket(self, ticket_ref: str) -> str:
        """Convierte referencias de ticket a formato TICKET_######."""
        if ticket_ref is None:
            return ""

        valor = str(ticket_ref).strip()
        if not valor:
            return ""

        if valor.isdigit():
            return f"TICKET_{int(valor):06d}"

        valor_upper = valor.upper()
        if valor_upper.startswith("TICKET_"):
            sufijo = valor_upper.split("_", 1)[1]
            if sufijo.isdigit():
                return f"TICKET_{int(sufijo):06d}"
            return valor_upper

        return valor_upper

    def actualizar_hora_salida_por_ticket(self, ticket_ref: str, hora_salida: Optional[str] = None) -> Dict:
        """
        Actualiza la hora de salida por número o código de ticket.

        Args:
            ticket_ref: Número (ej: 1) o código (ej: TICKET_000001)
            hora_salida: Hora de salida HH:MM:SS. Si no se envía, usa la hora actual.

        Returns:
            dict con resultado de la actualización.
        """
        try:
            codigo_ticket = self._normalizar_codigo_ticket(ticket_ref)
            if not codigo_ticket:
                return {
                    'success': False,
                    'error': 'Ticket inválido'
                }

            hora_final = hora_salida or datetime.now().strftime("%H:%M:%S")

            wb = openpyxl.load_workbook(self.ruta_excel)
            ws = wb["Registros"]

            for fila in range(2, ws.max_row + 1):
                numero_ticket_celda = str(ws.cell(row=fila, column=1).value or "").strip().upper()
                if numero_ticket_celda == codigo_ticket:
                    ws.cell(row=fila, column=6, value=hora_final)
                    wb.save(self.ruta_excel)
                    wb.close()
                    return {
                        'success': True,
                        'ticket': codigo_ticket,
                        'hora_salida': hora_final,
                        'mensaje': f'Hora de salida actualizada para {codigo_ticket}'
                    }

            wb.close()
            return {
                'success': False,
                'error': f'No se encontró el ticket {codigo_ticket}'
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def obtener_todos_tickets(self) -> Dict:
        """
        Lee el Excel maestro y retorna todos los tickets registrados.

        Returns:
            dict con estado y lista de tickets.
        """
        try:
            wb = openpyxl.load_workbook(self.ruta_excel, data_only=True)
            ws = wb["Registros"]

            tickets = []
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if not fila or not fila[0]:
                    continue

                hora_salida = fila[5]
                if hora_salida is None or str(hora_salida).strip() == "":
                    hora_salida = "No ha salido"

                tickets.append({
                    'numero_ticket': fila[0],
                    'nombres': fila[1] or "",
                    'apellidos': fila[2] or "",
                    'cedula': fila[3] or "",
                    'hora_ingreso': fila[4] or "",
                    'hora_salida': hora_salida,
                    'departamento': fila[6] or "",
                    'motivo': fila[7] or "",
                    'fecha_registro': fila[8] or "",
                })

            wb.close()
            return {
                'success': True,
                'total': len(tickets),
                'tickets': tickets,
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'total': 0,
                'tickets': [],
            }


# Instancia global (se inicializará en la API)
registro_manager = None


def inicializar_registro_manager(ruta_base: str):
    """Inicializa el manager de registros con una ruta base"""
    global registro_manager
    registro_manager = RegistroVehicular(ruta_base)
